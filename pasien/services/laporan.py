from collections import defaultdict
from datetime import date

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.dimensions import ColumnDimension, DimensionHolder
from openpyxl.worksheet.table import Table, TableStyleInfo
from pasien.services.helper import beautify_header
from django.db import connection
from django_pivot import pivot

# this entire file has bad codes from the worst depths of hell
# pls forgive


# def laporan_pendaftaran(tgl: date):
#     with connection.cursor() as cursor:
#         cursor.execute(
#             """
#         select
#             puskesmas_id as "PUSKESMAS",
#             rp.pulau as "PULAU",
#             SUM(case when penyakit_id = 'KATARAK' then 1 else 0 END) as "KATARAK",
#             SUM(case when penyakit_id = 'HERNIA' then 1 else 0 END) as "HERNIA",
#             SUM(case when penyakit_id = 'SUMBING' then 1 else 0 END) as "SUMBING",
#             SUM(case when penyakit_id = 'BENJOLAN' then 1 else 0 END) as "BENJOLAN",
#             COUNT(Id) as "TOTAL"
#         from pasien_pasien pp
#             inner join referensi_puskesmas rp on rp.puskesmas = pp.puskesmas_id
#         where pp.tanggal_nomor_antrian::date = date(%s)
#         group by puskesmas_id, rp.pulau
#         order by rp.pulau
#         """,
#             [tgl],
#         )
#         rows = cursor.fetchall()

#         report = []
#         footer_total = defaultdict(lambda: 0)
#         for row in rows:
#             report.append(
#                 {
#                     "PUSKESMAS": row[0],
#                     "PULAU": row[1],
#                     "KATARAK": row[2],
#                     "HERNIA": row[3],
#                     "SUMBING": row[4],
#                     "BENJOLAN": row[5],
#                     "TOTAL": row[6],
#                 }
#             )
#             footer_total[2] = footer_total[2] + int(row[2])
#             footer_total[3] = footer_total[3] + int(row[3])
#             footer_total[4] = footer_total[4] + int(row[4])
#             footer_total[5] = footer_total[5] + int(row[5])
#             footer_total[6] = footer_total[6] + int(row[6])

#         report.append(
#             {
#                 "PUSKESMAS": "TOTAL",
#                 "PULAU": "-",
#                 "KATARAK": footer_total[2],
#                 "HERNIA": footer_total[3],
#                 "SUMBING": footer_total[4],
#                 "BENJOLAN": footer_total[5],
#                 "TOTAL": footer_total[6],
#             }
#         )

#         return report


# def laporan_pendaftaran_excel():
#     full_report = laporan_pendaftaran()

#     colu = [row.pop("DAERAH") for row in full_report]

#     workbook = Workbook()
#     worksheet = workbook.active

#     worksheet["A1"] = "LAPORAN PENDAFTARAN"
#     worksheet["A1"].font = Font(size=20)

#     worksheet["A3"].value = "DAERAH"
#     for row, i in enumerate(colu):
#         column_cell = "A"
#         worksheet[column_cell + str(row + 4)] = str(i)

#     headers = full_report[0].keys()

#     for col, entry in enumerate(headers):
#         worksheet.cell(row=3, column=col + 2, value=beautify_header(entry))
#         worksheet.cell(row=3, column=col + 2).font = Font(bold=True)

#     row = 4
#     for data in full_report:
#         for col, entry in enumerate(data):
#             worksheet.cell(row=row, column=col + 2, value=data[entry])
#         row += 1

#     tab = Table(displayName="Table1", ref=f"A3:E{worksheet.max_row}")
#     style = TableStyleInfo(
#         name="TableStyleLight1",
#         showFirstColumn=False,
#         showLastColumn=False,
#         showRowStripes=False,
#         showColumnStripes=False,
#     )
#     tab.tableStyleInfo = style
#     worksheet.add_table(tab)

#     dim_holder = DimensionHolder(worksheet=worksheet)
#     for col in range(worksheet.min_column, worksheet.max_column + 1):
#         dim_holder[get_column_letter(col)] = ColumnDimension(
#             worksheet, min=col, max=col, width=20
#         )

#     worksheet.column_dimensions = dim_holder

#     return workbook


def laporan_kehadiran(tgl: date):
    with connection.cursor() as cursor:
        cursor.execute(
            """
        select 
            puskesmas_id as "PUSKESMAS",
            rp.pulau as "PULAU",
            SUM(case when penyakit_id = 'KATARAK' then 1 else 0 END) as "KATARAK",
            SUM(case when penyakit_id = 'HERNIA' then 1 else 0 END) as "HERNIA",
            SUM(case when penyakit_id = 'SUMBING' then 1 else 0 END) as "SUMBING",
            SUM(case when penyakit_id = 'BENJOLAN' then 1 else 0 END) as "BENJOLAN",
            COUNT(Id) as "TOTAL"
        from pasien_pasien pp 
            inner join referensi_puskesmas rp on rp.puskesmas = pp.puskesmas_id 
        where pp.tanggal_nomor_antrian::date = date(%s)
        group by puskesmas_id, rp.pulau
        order by rp.pulau 
        """,
            [tgl],
        )
        rows = cursor.fetchall()

        report = []
        footer_total = defaultdict(lambda: 0)
        for row in rows:
            report.append(
                {
                    "PUSKESMAS": row[0],
                    "PULAU": row[1],
                    "KATARAK": row[2],
                    "HERNIA": row[3],
                    "SUMBING": row[4],
                    "BENJOLAN": row[5],
                    "TOTAL": row[6],
                }
            )
            footer_total[2] = footer_total[2] + int(row[2])
            footer_total[3] = footer_total[3] + int(row[3])
            footer_total[4] = footer_total[4] + int(row[4])
            footer_total[5] = footer_total[5] + int(row[5])
            footer_total[6] = footer_total[6] + int(row[6])

        report.append(
            {
                "PUSKESMAS": "TOTAL",
                "PULAU": "-",
                "KATARAK": footer_total[2],
                "HERNIA": footer_total[3],
                "SUMBING": footer_total[4],
                "BENJOLAN": footer_total[5],
                "TOTAL": footer_total[6],
            }
        )

        return report


def laporan_kehadiran_excel():
    full_report = laporan_kehadiran()

    workbook = Workbook()
    worksheet = workbook.active

    penyakit_list = list(full_report.keys())

    worksheet["A1"] = "LAPORAN KEHADIRAN"
    worksheet["A1"].font = Font(size=20)

    worksheet["A3"].value = "JENIS PENYAKIT"
    for row, i in enumerate(penyakit_list):
        column_cell = "A"
        worksheet[column_cell + str(row + 4)] = str(i)

    header_list = list(full_report[penyakit_list[0]].keys())

    for col, entry in enumerate(header_list):
        worksheet.cell(row=3, column=col + 2, value=beautify_header(entry))
        worksheet.cell(row=3, column=col + 2).font = Font(bold=True)

    row = 4
    for penyakit in penyakit_list:
        penyakit_data = list(full_report[penyakit].values())
        for col, entry in enumerate(penyakit_data):
            worksheet.cell(row=row, column=col + 2, value=entry)
        row += 1

    dim_holder = DimensionHolder(worksheet=worksheet)
    for col in range(worksheet.min_column, worksheet.max_column + 1):
        dim_holder[get_column_letter(col)] = ColumnDimension(
            worksheet, min=col, max=col, width=20
        )

    worksheet.column_dimensions = dim_holder

    tab = Table(displayName="Table1", ref=f"A3:M{worksheet.max_row}")
    style = TableStyleInfo(
        name="TableStyleLight1",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=False,
        showColumnStripes=False,
    )
    tab.tableStyleInfo = style
    worksheet.add_table(tab)

    return workbook


def laporan_screening(tgl: date):
    with connection.cursor() as cursor:
        cursor.execute(
            """
        select pp.diagnosa as "PENYAKIT",
            count(*) as "TOTAL_PASIEN",
            sum(case when pp.tanggal_nomor_antrian::date = date(%s) then 1 else 0 end) as "TOTAL_HADIR",
            sum(case when pp.perlu_rescreen = true and pp.tanggal_nomor_antrian::date = date(%s) then 1 else 0 end) as "PERLU_RESCREEN",
            sum(case when ps.telah_lewat_cek_tensi = true and ps.jam_cek_tensi::date = date(%s) then 1 else 0 end) as "HADIR_TENSI",
            sum(case when ps.telah_lewat_cek_tensi = false and ps.jam_cek_tensi::date = date(%s) then 1 else 0 end) as "GAGAL_TENSI",
            sum(case when ps.telah_lewat_pemeriksaan = true and ps.jam_pemeriksaan::date = date(%s) then 1 else 0 end) as "HADIR_PEMERIKSAAN",
            sum(case when ps.telah_lewat_pemeriksaan = false and ps.jam_pemeriksaan::date = date(%s) then 1 else 0 end) as "GAGAL_PEMERIKSAAN",
            sum(case when ps.telah_lewat_cek_lab = true and ps.jam_cek_lab::date = date(%s) then 1 else 0 end) as "HADIR_LAB",
            sum(case when ps.telah_lewat_cek_radiologi = true and ps.jam_cek_radiologi::date = date(%s) then 1 else 0 end) as "HADIR_RONTGEN",
            sum(case when ps.telah_lewat_cek_ekg  = true and ps.jam_cek_ekg::date = date(%s) then 1 else 0 end) as "HADIR_EKG",
            sum(case when ps.telah_lewat_cek_kartu_kuning = true and ps.jam_cek_kartu_kuning::date = date(%s) then 1 else 0 end) as "HADIR_KARTUKUNING",
            sum(case when pk.status = 'LOLOS' and pk.waktu_dibuat::date = date(%s) then 1 else 0 end) as "KARTUKUNING_LOLOS",
            sum(case when pk.status = 'GAGAL' and pk.waktu_dibuat::date = date(%s) then 1 else 0 end) as "KARTUKUNING_GAGAL",
            sum(case when pk.status = 'PENDING' and pk.waktu_dibuat::date = date(%s) then 1 else 0 end) as "KARTUKUNING_PENDING"
        from pasien_pasien pp 
            left join pasien_screeningpasien ps on ps.pasien_id = pp.id
            left join pasien_kartukuning pk on pk.pasien_id = pp.id
        group by pp.diagnosa
        order by diagnosa 
        """,
            [tgl, tgl, tgl, tgl, tgl, tgl, tgl, tgl, tgl, tgl, tgl, tgl, tgl],
        )
        rows = cursor.fetchall()

        raw_data = []
        for row in rows:
            raw_data.append(
                {
                    "PENYAKIT": row[0],
                    "TOTAL_PASIEN": row[1],
                    "TOTAL_HADIR": row[2],
                    "PERLU_RESCREEN": row[3],
                    "HADIR_TENSI": row[4],
                    "GAGAL_TENSI": row[5],
                    "HADIR_PEMERIKSAAN": row[6],
                    "GAGAL_PEMERIKSAAN": row[7],
                    "HADIR_LAB": row[8],
                    "HADIR_RONTGEN": row[9],
                    "HADIR_EKG": row[10],
                    "HADIR_KARTUKUNING": row[11],
                    "KARTUKUNING_LOLOS": row[12],
                    "KARTUKUNING_GAGAL": row[13],
                    "KARTUKUNING_PENDING": row[14],
                }
            )

        dict_data = {}
        for row in rows:
            dict_data[row[0]] = row

        row_headers = [
            "TOTAL_PASIEN",
            "TOTAL_HADIR",
            "PERLU_RESCREEN",
            "HADIR_TENSI",
            "GAGAL_TENSI",
            "HADIR_PEMERIKSAAN",
            "GAGAL_PEMERIKSAAN",
            "HADIR_LAB",
            "HADIR_RONTGEN",
            "HADIR_EKG",
            "HADIR_KARTUKUNING",
            "KARTUKUNING_LOLOS",
            "KARTUKUNING_GAGAL",
            "KARTUKUNING_PENDING",
        ]
        i = 1
        pivot_data = []

        for row_header in row_headers:
            pivot_data.append(
                {
                    "SCREENING": row_header,
                    "KATARAK": dict_data["KATARAK"][i],
                    "PTERYGIUM": dict_data["PTERYGIUM"][i],
                    "HERNIA": dict_data["HERNIA"][i],
                    "SUMBING": dict_data["SUMBING"][i],
                    "BENJOLAN": dict_data["BENJOLAN"][i],
                    "MINOR GA": dict_data["MINOR GA"][i],
                    "MINOR LOKAL": dict_data["MINOR LOKAL"][i],
                    "TOTAL": dict_data["KATARAK"][i]
                    + dict_data["PTERYGIUM"][i]
                    + dict_data["HERNIA"][i]
                    + dict_data["SUMBING"][i]
                    + dict_data["BENJOLAN"][i]
                    + dict_data["MINOR GA"][i]
                    + dict_data["MINOR LOKAL"][i],
                }
            )
            i = i + 1

        return pivot_data


def laporan_screening_excel():
    full_report = laporan_screening()

    workbook = Workbook()
    worksheet = workbook.active

    penyakit_list = list(full_report.keys())

    worksheet["A1"] = "LAPORAN SCREENING"
    worksheet["A1"].font = Font(size=20)

    worksheet["A3"].value = "JENIS PENYAKIT"
    for row, i in enumerate(penyakit_list):
        column_cell = "A"
        worksheet[column_cell + str(row + 4)] = str(i)

    header_list = list(full_report[penyakit_list[0]].keys())

    for col, entry in enumerate(header_list):
        worksheet.cell(row=3, column=col + 2, value=beautify_header(entry))
        worksheet.cell(row=3, column=col + 2).font = Font(bold=True)

    row = 4
    for penyakit in penyakit_list:
        penyakit_data = list(full_report[penyakit].values())
        for col, entry in enumerate(penyakit_data):
            worksheet.cell(row=row, column=col + 2, value=entry)
        row += 1

    dim_holder = DimensionHolder(worksheet=worksheet)
    for col in range(worksheet.min_column, worksheet.max_column + 1):
        dim_holder[get_column_letter(col)] = ColumnDimension(
            worksheet, min=col, max=col, width=20
        )

    worksheet.column_dimensions = dim_holder

    tab = Table(displayName="Table1", ref=f"A3:M{worksheet.max_row}")
    style = TableStyleInfo(
        name="TableStyleLight1",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=False,
        showColumnStripes=False,
    )
    tab.tableStyleInfo = style
    worksheet.add_table(tab)

    return workbook
