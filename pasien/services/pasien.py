from datetime import date, datetime, time, tzinfo
from doctest import master

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.dimensions import ColumnDimension, DimensionHolder
from openpyxl.worksheet.table import Table, TableStyleInfo

from pasien.models import KartuKuning, Pasien
from referensi.models import Penyakit, Puskesmas

from django.core.exceptions import ValidationError
from django.utils import timezone
from pasien.services.helper import beautify_header


def generate_import_template():
    workbook = Workbook()
    worksheet = workbook.active

    worksheet["A1"] = "Template Import Pasien"
    worksheet["A1"].font = Font(size=20)

    headers = [
        "Nomor Seri",
        "Puskesmas",
        "Penyakit",
        "Nama",
        "Nomor Identitas",
        "Tipe Identitas",
        "Tempat Lahir",
        "Tanggal Lahir (DD-MM-YYYY)",
        "Jenis Kelamin (L/P)",
        "Alamat",
        "Kabupaten/Kota",
        "Nomor Telepon",
        "Nama Keluarga",
        "Nomor Telepon Keluarga",
    ]

    worksheet.move_range("A2:N2", rows=1, cols=1)
    worksheet.append(headers)

    dim_holder = DimensionHolder(worksheet=worksheet)

    for col in range(worksheet.min_column, worksheet.max_column + 1):
        dim_holder[get_column_letter(col)] = ColumnDimension(
            worksheet, min=col, max=col, width=20
        )

    worksheet.column_dimensions = dim_holder

    tab = Table(displayName="Table1", ref="A3:N4")
    style = TableStyleInfo(
        name="TableStyleLight8",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=True,
    )
    tab.tableStyleInfo = style
    worksheet.add_table(tab)

    puskesmas_list = Puskesmas.objects.all().order_by("puskesmas")
    penyakit_list = Penyakit.objects.all().order_by("nama")
    tipe_identitas = ["KK", "KTP", "PASPOR", "SIM"]

    master_data_worksheet = workbook.create_sheet("Master Data")

    master_data_worksheet["A1"].value = "Puskesmas"
    master_data_worksheet["A1"].font = Font(bold=True)

    master_data_worksheet["C1"].value = "Penyakit"
    master_data_worksheet["C1"].font = Font(bold=True)

    master_data_worksheet["E1"].value = "Tipe Identitas"
    master_data_worksheet["E1"].font = Font(bold=True)

    master_data_worksheet["G1"].value = "Kabupaten/Kota"
    master_data_worksheet["G1"].font = Font(bold=True)

    for row, i in enumerate(puskesmas_list):
        column_cell = "A"
        master_data_worksheet[column_cell + str(row + 2)] = i.puskesmas

    for row, i in enumerate(penyakit_list):
        column_cell = "C"
        master_data_worksheet[column_cell + str(row + 2)] = i.nama

    for row, i in enumerate(tipe_identitas):
        column_cell = "E"
        master_data_worksheet[column_cell + str(row + 2)] = i

    for row, i in enumerate(
        puskesmas_list.order_by("pulau").values_list("pulau", flat=True).distinct()
    ):
        column_cell = "G"
        master_data_worksheet[column_cell + str(row + 2)] = i

    dim_holder = DimensionHolder(worksheet=master_data_worksheet)

    for col in range(
        master_data_worksheet.min_column, master_data_worksheet.max_column + 1
    ):
        dim_holder[get_column_letter(col)] = ColumnDimension(
            master_data_worksheet, min=col, max=col, width=20
        )

    return workbook


def import_pasien(file):
    worksheet = load_workbook(file)

    worksheet = worksheet.active

    if worksheet["A1"].value != "Template Import Pasien":
        raise ValidationError("Invalid file, please check again!")

    puskesmas_list = list(Puskesmas.objects.all())
    penyakit_list = list(Penyakit.objects.all())

    now = datetime.now()

    new_patients = []
    for row in worksheet.iter_rows(
        min_row=4, min_col=1, max_row=worksheet.max_row, max_col=worksheet.max_column
    ):
        # count years
        date_of_birth = now - row[7].value
        years = date_of_birth.total_seconds() / (365.242 * 24 * 3600)
        tahun = int(years)
        months = (years - tahun) * 12
        bulan = int(months)
        umur = f"{tahun} Tahun {bulan} Bulan"

        inputted_puskemas = str(row[1].value).upper().strip()
        inputted_penyakit = str(row[2].value).upper().strip()
        inputted_daerah = str(row[10].value).upper().strip()

        if (
            next(
                filter(lambda x: x.puskesmas == inputted_puskemas, puskesmas_list), None
            )
            is None
        ):
            raise ValidationError(
                f"Row: {row[0].row} - Data Puskesmas yang ditaruh tidak valid."
            )

        if (
            next(filter(lambda x: x.nama == inputted_penyakit, penyakit_list), None)
            is None
        ):
            raise ValidationError(
                f"Row: {row[0].row} - Data Penyakit yang ditaruh tidak valid."
            )

        if (
            next(filter(lambda x: x.pulau == inputted_daerah, puskesmas_list), None)
            is None
        ):
            raise ValidationError(
                f"Row: {row[0].row} - Daerah yang ditaruh tidak valid."
            )

        puskesmas = next(
            filter(lambda x: x.puskesmas == inputted_puskemas, puskesmas_list)
        )
        penyakit = next(filter(lambda x: x.nama == inputted_penyakit, penyakit_list))

        new_patients.append(
            Pasien(
                nomor_seri=row[0].value,
                puskesmas=puskesmas,
                penyakit=penyakit,
                nama=row[3].value,
                nomor_identitas=row[4].value,
                tipe_identitas=str(row[5].value).upper().strip(),
                tempat_lahir=row[6].value,
                tanggal_lahir=row[7].value,
                umur=umur,
                jenis_kelamin=row[8].value,
                alamat=row[9].value,
                daerah=inputted_daerah,
                pulau=puskesmas.pulau,
                nomor_telepon=row[11].value,
                nama_keluarga=row[12].value,
                nomor_telepon_keluarga=row[13].value,
                diagnosa=penyakit.pk,
                nama_pendamping=row[12].value,
                nomor_telepon_pendamping=row[13].value,
            )
        )

        __validate_input_data(new_patients)

    Pasien.objects.bulk_create(new_patients)


def serah_nomor_antrian(pasien: Pasien, nomor_antrian: str):
    if pasien.perlu_rescreen:
        pasien.nomor_antrian_pertama = pasien.nomor_antrian
        pasien.tanggal_nomor_antrian_pertama = pasien.tanggal_nomor_antrian
    pasien.nomor_antrian = nomor_antrian
    pasien.tanggal_nomor_antrian = timezone.now()
    pasien.last_status = "DAFTAR"
    pasien.save()


def batal_nomor_antrian(pasien: Pasien):
    today = timezone.now()
    if pasien.tanggal_nomor_antrian.date() != today.date():
        raise ValidationError("Pembatalan nomor antrian harus di hari yang sama!")
    pasien.nomor_antrian = None
    pasien.tanggal_nomor_antrian = None
    pasien.last_status = None
    pasien.save()


def update_penyakit(pasien: Pasien, penyakit: str):
    pasien.penyakit_id = penyakit
    pasien.diagnosa = penyakit
    pasien.save()


def update_diagnosa(pasien: Pasien, diagnosa: str):
    pasien.diagnosa = diagnosa
    pasien.save()


def pending_tensi(pasien: Pasien):
    pasien.perlu_rescreen = True
    pasien.save()


def serah_kartu_kuning(
    pasien: Pasien, status: str, tanggal: date, jam: time, perhatian: list
):
    kartu_kuning: KartuKuning = KartuKuning()
    kartu_kuning.nomor = __generate_nomor_kartu_kuning()
    kartu_kuning.pasien = pasien
    kartu_kuning.jam_operasi = jam
    kartu_kuning.tanggal_operasi = tanggal
    kartu_kuning.perhatian = perhatian
    kartu_kuning.status = status
    kartu_kuning.save()


def batal_serah_kartu_kuning(pasien: Pasien):
    KartuKuning.objects.filter(pasien=pasien).delete()


def __generate_nomor_kartu_kuning():
    prefix = "KK"
    running_no = 1
    last_record = KartuKuning.objects.all().order_by("-nomor").first()

    if last_record:
        last_running_no = int(last_record.nomor[-4:])
        running_no = running_no + last_running_no

    return prefix + str(running_no).zfill(4)


def __validate_input_data(pasien_list: list[Pasien]):
    NON_NULL_VALUES = [
        "nomor_seri",
        "nama",
        "nomor_identitas",
        "tempat_lahir",
        "jenis_kelamin",
        "alamat",
        "tipe_identitas",
        "tanggal_lahir",
        "nomor_telepon",
        "nama_keluarga",
        "nomor_telepon_keluarga",
    ]
    DEFAULT_GENDER_VALUE = ["L", "P"]
    CHECK_MAX_LENGTH_VALUE = [
        "nomor_identitas",
        "nomor_telepon",
        "nomor_telepon_keluarga",
    ]
    all_pasien: list[Pasien] = list(Pasien.objects.all())
    for row, pasien in enumerate(pasien_list, 1):
        for non_null_value in NON_NULL_VALUES:
            value = getattr(pasien, non_null_value)
            if value is None:
                raise ValidationError(
                    f"Row: {row + 3} - {beautify_header(non_null_value)} tidak ditaruh."
                )

        for check_value in CHECK_MAX_LENGTH_VALUE:
            value = str(getattr(pasien, check_value))
            if len(value) > 30:
                raise ValidationError(
                    f"Row: {row + 3} - {beautify_header(non_null_value)} yang ditaruh lebih dari 30 karakter."
                )

        if next(filter(lambda x: x.nomor_seri == pasien.nomor_seri, all_pasien), None):
            raise ValidationError(
                f"Row: {row + 3} - No. Seri ({pasien.nomor_seri}) yang ditaruh telah terdaftar."
            )

        if next(
            filter(lambda x: x.nomor_identitas == pasien.nomor_identitas, all_pasien),
            None,
        ):
            raise ValidationError(
                f"Row: {row + 3} - No. Identitas ({pasien.nomor_identitas}) yang ditaruh telah terdaftar."
            )

        if pasien.jenis_kelamin not in DEFAULT_GENDER_VALUE:
            raise ValidationError(
                f"Row: {row + 3} - Data Jenis Kelamin yang ditaruh tidak valid."
            )
